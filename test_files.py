import codecs
import pypdf
import csv
from zipfile import ZipFile
from openpyxl import load_workbook


def test_pdf():
    with ZipFile("resources/difiles.zip") as myzip:
        with myzip.open("tut.pdf") as myfile:
            reader = pypdf.PdfReader(myfile)
            pageread = (reader.pages[2].extract_text())
            print(pageread)
    assert "Python è un linguaggio" in pageread

def test_xlsx():
    with ZipFile("resources/difiles.zip") as myzip:
        with myzip.open("file_example_XLSX_50.xlsx") as myfile:
            workbook = load_workbook(myfile)
            sheet = workbook.active
            country = (sheet.cell(3, 5).value)
            print(country)
    assert country == "Great Britain"


def test_csv():
    with ZipFile("resources/difiles.zip") as myzip:
        with myzip.open("addresses.csv", 'r') as myfile:
            reader = csv.DictReader(codecs.iterdecode(myfile, 'utf-8'))
            line = (next(reader))
            print(line)
    assert [row for row in line] == ['John', 'Doe', '120 jefferson st.', 'Riverside', ' NJ', ' 08075']
